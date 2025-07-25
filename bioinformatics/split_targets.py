import io
import os

import openpyxl
from nicegui import ui
import pandas as pd

from common.download import download
from log import PrintLog


def split_molecule_targets(v,l = PrintLog()):
    try:
        l.info(v)
        wb = openpyxl.load_workbook(io.BytesIO(v.content.read()), read_only=True)
        # 获取所有sheet名
        sheet_names = wb.sheetnames

        buf = io.BytesIO()
        with pd.ExcelWriter(buf) as writer:
            for name in sheet_names:
                ws = wb[name]
                data = []
                for row in ws.iter_rows(values_only=True):
                    if row[0].strip() == '#Name' or row[0].strip() == '':
                        continue
                    mole = row[0]
                    for item in row[2:]:
                        if item is not None and item.strip() != '':
                            item = item.split('|')[1]
                            item = item.split('(')[0]
                            data.append({
                                'name': mole,
                                'target': item
                            })
                df = pd.DataFrame(data)
                df.to_excel(writer, sheet_name=name, index=False)
        download(buf.getvalue(), filename=f"{v.name}_split.xlsx",media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        l.success("下载成功")
    except Exception as e:
        l.error(f"发生异常：{e}")
    finally:
        ui.timer(15.0, lambda: os.unlink(f"{v.name}_split.xlsx") if os.path.exists(f"{v.name}_split.xlsx") else None, once=True)


